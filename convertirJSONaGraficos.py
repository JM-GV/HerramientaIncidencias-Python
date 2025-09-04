import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
import locale

# Configurar el idioma a español para los nombres de los meses
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')  # Cambiar a 'es_ES' según tu sistema operativo

def convertirJSONaGraficos(json_file, excel_file):
    # Cargar el archivo JSON
    with open(json_file, 'r') as file:
        data = json.load(file)

    # Crear un DataFrame principal para las incidencias
    incidencias = []
    for item in data:
        incidencia = {
            "ID": item["id"],
            "Título": item["titulo"],
            "Descripción": item["descripcion"],
            "Prioridad": item["prioridad"],
            "Asignado a": item["asignado_a"],
            "Estado": item["estado"],
            "Creado en": item["creado_en"]
        }
        incidencias.append(incidencia)

    df_incidencias = pd.DataFrame(incidencias)

    # Crear el archivo Excel
    wb = Workbook()

    # Hoja 1: Información Incidencias
    ws_info = wb.active
    ws_info.title = "Información Incidencias"

    # Escribir los datos del DataFrame en la hoja
    for r in dataframe_to_rows(df_incidencias, index=False, header=True):
        ws_info.append(r)

    # Hoja 2: G. Estado Incidencias
    ws_estado = wb.create_sheet(title="G. Estado Incidencias")

    # Contar incidencias por estado
    incidencias_por_estado = df_incidencias["Estado"].value_counts()

    # Escribir los datos en la hoja
    ws_estado.append(["Estado", "Número de Incidencias"])
    for estado, cantidad in incidencias_por_estado.items():
        ws_estado.append([estado, cantidad])

    # Crear el gráfico de pastel
    pie = PieChart()
    labels = Reference(ws_estado, min_col=1, min_row=2, max_row=1 + len(incidencias_por_estado))
    data = Reference(ws_estado, min_col=2, min_row=2, max_row=1 + len(incidencias_por_estado))
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Estado de las Incidencias"
    ws_estado.add_chart(pie, "E5")

    # Hoja 3: G. Carga de Trabajo
    ws_carga = wb.create_sheet(title="G. Carga de Trabajo")

    # Contar incidencias por usuario (excluyendo las finalizadas)
    df_incidencias_no_finalizado = df_incidencias[df_incidencias["Estado"] != "Finalizado"]
    incidencias_por_usuario = df_incidencias_no_finalizado["Asignado a"].value_counts()

    # Escribir los datos en la hoja
    ws_carga.append(["Usuario", "Número de Incidencias"])
    for usuario, cantidad in incidencias_por_usuario.items():
        ws_carga.append([usuario, cantidad])

    # Crear el gráfico de barras
    bar = BarChart()
    labels = Reference(ws_carga, min_col=1, min_row=2, max_row=1 + len(incidencias_por_usuario))
    data = Reference(ws_carga, min_col=2, min_row=2, max_row=1 + len(incidencias_por_usuario))
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(labels)
    bar.title = "Carga de Trabajo por Usuario"
    bar.x_axis.title = "Usuario"
    bar.y_axis.title = "Número de Incidencias"
    ws_carga.add_chart(bar, "E5")

    # Hoja 4: G. Incidencias por Mes
    ws_mes = wb.create_sheet(title="G. Incidencias por Mes")

    # Contar incidencias por mes
    df_incidencias["Mes"] = pd.to_datetime(df_incidencias["Creado en"]).dt.strftime('%B')
    incidencias_por_mes = df_incidencias["Mes"].value_counts().sort_index()

    # Escribir los datos en la hoja
    ws_mes.append(["Mes", "Número de Incidencias"])
    for mes, cantidad in incidencias_por_mes.items():
        ws_mes.append([mes, cantidad])

    # Crear el gráfico de pastel
    pie_mes = PieChart()
    labels = Reference(ws_mes, min_col=1, min_row=2, max_row=1 + len(incidencias_por_mes))
    data = Reference(ws_mes, min_col=2, min_row=2, max_row=1 + len(incidencias_por_mes))
    pie_mes.add_data(data, titles_from_data=False)
    pie_mes.set_categories(labels)
    pie_mes.title = "Incidencias por Mes"
    ws_mes.add_chart(pie_mes, "E5")

    # Guardar el archivo Excel
    wb.save(excel_file)
    print(f"Archivo Excel '{excel_file}' generado exitosamente.")

# Ejemplo de uso
convertirJSONaGraficos("incidencias.json", "incidencias.xlsx")
